# encoding:utf-8
import re
import os
import sys
import shutil
import time
import openpyxl

from common import output_info_redirect


def chk_proportion_file(path, wt_logfile):
    output_info_redirect(u'开始检查%s...\n' % path, wt_logfile)
    filename = u'费用占比.xlsx'
    file_path = path + '\\' + filename
    if not os.path.exists(path):
        output_info_redirect(u'未找到%s, 将自动创建改目录...\n' % path, wt_logfile)
        os.mkdir(path)
    if os.path.isfile(file_path):
        now_time = time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime(time.time()))
        file_path_tmp = file_path.split('.xlsx')[0]
        file_path_back = file_path_tmp + '_backup_at_' + now_time + '.xlsx'
        output_info_redirect(u'发现 费用占比.xlsx 文件已存在, 将重命名原文件...\n', wt_logfile)
        shutil.move(file_path, file_path_back)
    output_info_redirect(u'检查完毕!\n', wt_logfile)
    return file_path


def calculate_cost_and_income(file_path, wt_logfile):
    filename = file_path.split('\\')[-1]
    output_info_redirect(u'开始统计 %s 的费用占比...\n' % filename, wt_logfile)
    items_and_code = {'outsource_repair_cost': ['6401.01.001.0003-', '6401.01.002.0004-', '6401.01.003.0004-'],
                      'human_cost': ['6401.99-'],
                      'fitting_cost': ['6401.01.002.0002-', '6401.01.003.0002-', '6401.02.001.0002-'],
                      'petrol_cost': ['6401.88.001-', '6401.04.002.0002-'],
                      'road_toll': ['6401.88.002-', '6401.04.002.0003-'],
                      'parking_fee': ['6401.88.003-', '6401.04.002.0004-'],
                      'main_business_cost': ['6401-'],
                      'main_business_income': ['6001.01-', '6001.02-']}
    code_and_value = {'6401.01.001.0003-': 0, '6401.01.002.0004-': 0, '6401.01.003.0004-': 0, '6401.99-': 0,
                      '6401.01.002.0002-': 0, '6401.01.003.0002-': 0, '6401.02.001.0002-': 0,
                      '6401.88.001-': 0, '6401.04.002.0002-': 0, '6401.88.002-': 0, '6401.04.002.0003-': 0,
                      '6401.88.003-': 0, '6401.04.002.0004-': 0, '6401-': 0, '6001.01-': 0, '6001.02-': 0}
    items_and_cost = {}

    f = open(file_path)
    content_list = f.read().split('\n')
    f.close()

    shop_name = file_path.split(u'店')[0].split('\\')[-1] + u'店'
    shop_code = re.split('[>-]', content_list[95])[1]
    items_and_cost['shop_name'] = shop_name
    items_and_cost['shop_code'] = shop_code
    output_info_redirect(u'%s 的门店编号是: %s\n' % (shop_name, shop_code), wt_logfile)

    for code in code_and_value:
        for line in content_list:
            if code in line:
                line_num = content_list.index(line) + 1
                if content_list[line_num] == '</TD>':
                    line_num += 1
                code_and_value[code] = float(re.split('[=>]', content_list[line_num])[2])
                break
    output_info_redirect(u'%s 获取的各项数据为: %s\n' % (shop_name, code_and_value), wt_logfile)

    for item in items_and_code:
        item_sum = 0
        for code in items_and_code[item]:
            item_sum += code_and_value[code]
        items_and_cost[item] = item_sum
    output_info_redirect(u'%s 的费用占比数据为: %s\n' % (shop_name, items_and_cost), wt_logfile)

    return shop_name, items_and_cost


def merge_proportion_result(proportion_file, src_path, wt_logfile):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = u'费用占比'
    line1 = [u'门店代码', u'门店名称', u'外包修理费', u'人员成本', u'配件成本', u'人员、配件成本合计', u'外包修理费占比成本',
             u'外包修理费占比收入', u'汽油费', u'路桥费', u'停车费', u'巡检费合计', u'主营业务成本', u'巡检费占比成本百分比',
             u'主营业务收入(不含暂估)', u'巡检费占收入百分比']
    for i in range(len(line1)):
        col = i + 1
        ws.cell(1, col, line1[i])

    raw = 2
    for shop in os.listdir(src_path):
        shop_file_path = src_path + '\\' + shop
        name_and_values = calculate_cost_and_income(shop_file_path, wt_logfile)
        shop_name = name_and_values[0]
        shop_values = name_and_values[1]
        output_info_redirect(u'正在将 %s 的数据写入 费用占比.xlsx 中...\n' % shop_name, wt_logfile)
        ws.cell(raw, 1, shop_values['shop_code'])
        ws.cell(raw, 2, shop_values['shop_name'])
        ws.cell(raw, 3, shop_values['outsource_repair_cost'])
        ws.cell(raw, 4, shop_values['human_cost'])
        ws.cell(raw, 5, shop_values['fitting_cost'])
        ws.cell(raw, 6, "=D%s+E%s" % (raw, raw))
        ws.cell(raw, 7, "=C%s/F%s" % (raw, raw))
        ws.cell(raw, 8, "=C%s/O%s" % (raw, raw))
        ws.cell(raw, 9, shop_values['petrol_cost'])
        ws.cell(raw, 10, shop_values['road_toll'])
        ws.cell(raw, 11, shop_values['parking_fee'])
        ws.cell(raw, 12, "=I%s+J%s+K%s" % (raw, raw, raw))
        ws.cell(raw, 13, shop_values['main_business_cost'])
        ws.cell(raw, 14, "=L%s/M%s" % (raw, raw))
        ws.cell(raw, 15, shop_values['main_business_income'])
        ws.cell(raw, 16, "=L%s/O%s" % (raw, raw))
        output_info_redirect(u'写入完毕!\n', wt_logfile)
        raw += 1

    wb.save(proportion_file)
    output_info_redirect(u'保存 费用占比.xlsx 完成!\n', wt_logfile)
    text = u'###' * 20 + \
           u'\n请按 [回车键] 退出...'
    output_info_redirect(text, wt_logfile)
    raw_input()
    time.sleep(1)
    return

