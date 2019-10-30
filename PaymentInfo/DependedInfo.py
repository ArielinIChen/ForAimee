# encoding:utf-8
import os
import sys
import time
from openpyxl import load_workbook, Workbook
from common import output_info_redirect


def shop_depart_relationship(bank_info_xl_path, logfile):
    # bank_info_xl_path = u'E:\诺互银行\银行信息资料\制单银行-门店对应表.xlsx'
    sheet_name = u'有银行的'
    relation_wb = load_workbook(bank_info_xl_path)
    if sheet_name in relation_wb.sheetnames:
        relation_sheet = relation_wb[sheet_name]
        relation_dict = {}
        # key 部门名称, 如: 总经办, 华东一区, 成都XXXX综合店;
        #  value 制单部门, 如: 上海诺互制单, 成都八分制单

        for i in range(2, relation_sheet.max_row + 1):
            shop_name = relation_sheet.cell(i, 2).value
            depart_name = relation_sheet.cell(i, 3).value
            if shop_name and depart_name:
                relation_dict[shop_name] = depart_name
            else:
                continue
        return relation_dict
    else:
        text = u'%s 没有找到 [%s]sheet, 请检查... \n' \
               u'系统将在3秒后退出 \n' % (bank_info_xl_path, sheet_name)
        output_info_redirect(text, logfile)
        time.sleep(3)
        sys.exit(1)


def get_payee_info(payee_info_xl_path):
    # return data looks like {收款人账号1: {最终名称1, 开户银行1, 省份1, 市1}, ......}
    # payee_info_xl_path = u'E:\诺互银行\银行制单模板\打款资料库.xlsx'
    sheet_name = '123'
    payee_wb = load_workbook(payee_info_xl_path)
    payee_sheet = payee_wb[sheet_name]
    payee_dict = {}

    for i in range(2, payee_sheet.max_row + 1):
        account = payee_sheet.cell(i, 1).value
        name = payee_sheet.cell(i, 2).value
        bank = payee_sheet.cell(i, 3).value
        province = payee_sheet.cell(i, 4).value
        city = payee_sheet.cell(i, 5).value
        payee_dict[account] = {'name': name, 'bank': bank, 'province': province, 'city': city}

    return payee_dict


def get_paying_bank_info(bank_check_xl_path):
    # return data looks like {记账部门1: [付款账号1, 地区A], 记账部门2: [付款账号2, 地区B]......}
    # bankcheck_xl_path = u'E:\诺互银行\银行核对\银行核对公式.xlsx'
    sheet_name = u'银行明细'
    bankcheck_wb = load_workbook(bank_check_xl_path)
    paying_bank_detail_sheet = bankcheck_wb[sheet_name]
    paying_bank_detail_dict = {}

    for i in range(2, paying_bank_detail_sheet.max_row + 1):
        _making_depart = paying_bank_detail_sheet.cell(i, 13).value
        if _making_depart:
            region = paying_bank_detail_sheet.cell(i, 14).value
            paying_account = paying_bank_detail_sheet.cell(i, 3).value
            paying_bank_detail_dict[_making_depart] = [paying_account, region]

    return paying_bank_detail_dict


if __name__ == '__main__':
    print get_payee_info(u'E:\诺互银行\银行制单模板\打款资料库.xlsx')
