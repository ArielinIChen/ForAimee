# encoding:utf-8
from openpyxl import load_workbook
from openpyxl.styles import numbers

# ERROR_SHEET格式: 总表的行号, 分表的行号, 错误分类, 总表收方名称, 总表付款金额, 总表记账部门, 分表付方账号(仅在付方账号有误时记录)
#
# 1 从[上传模板sheet]中逐行取出不为空的数据, 记入ori_data字典中; key为: 行号, value为: [收款人, 金额, 记账部门]
# 2 从[银行明细sheet]中逐行取出不为空的数据, 记入bank_info字典中; key为: 银行账号, value为: 记账部门
# 3 从[银行导出sheet]中逐行取出每一行数据, '行号', '付方账号', '收款人', '金额', '记账部门'
#     以'付方账号'为标准, 向bank_info搜索'记账部门';
#         如果不一致, 将此条数据标记为: "记账部门有误"; 并计入output_data
#         如果一致, 直接计入output_data
# 4 遍历ori_data, 向output_data做匹配
#     如果匹配得到, 并且output_data没有 [记账部门有误] ; pass
#     如果匹配得到, 并且output_data有 [记账部门有误] ; 记录两边行号, 写入Error_Sheet
#     如果匹配不到, 此条目标记为: "未匹配成功" ; 记录两边行号, 写入Error_Sheet


def payment_check(filepath):
    oriName = u'上传模板'
    bankDetail = u'银行明细'
    bankOutput = u'银行导出'
    wb = load_workbook(filepath)
    ori_sheet = wb[oriName]
    bank_detail_sheet = wb[bankDetail]
    output_sheet = wb[bankOutput]
    bank_info = {}
    output_data = {}

    sheet_num = len(wb.worksheets)
    err_sheet = wb.create_sheet(u'错误的数据', sheet_num + 1)
    err_sheet.cell(1, 1, '总表的行号')
    err_sheet.cell(1, 2, '分表的行号')
    err_sheet.cell(1, 3, '错误分类')
    err_sheet.cell(1, 4, '总表收方名称')
    err_sheet.cell(1, 5, '总表付款金额')
    err_sheet.cell(1, 6, '总表记账部门')
    err_sheet.cell(1, 7, '分表付方账号(仅在付方账号有误时记录)')
    err_sheet_row = 2

    # 遍历[银行明细sheet], 当 "银行账号" 和 "记账部门" 都不为空时, 记入bank_info字典
    for i in range(2, bank_detail_sheet.max_row + 1):
        _bank_account = bank_detail_sheet.cell(i, 3).value
        _depart_name = bank_detail_sheet.cell(i, 13).value

        if _depart_name is None or _bank_account is None:
            continue

        _bank_account = _bank_account.strip()
        _depart_name = _depart_name.strip()
        bank_info[_bank_account] = _depart_name

    #
    for i in range(2, output_sheet.max_row + 1):
        _op_account = output_sheet.cell(i, 4).value.split(u',')[1]
        _op_money = output_sheet.cell(i, 5).value
        _op_user = output_sheet.cell(i, 6).value
        _op_depart = output_sheet.cell(i, 9).value

        if _op_account is not None:
            _op_account = _op_account.strip()

        if _op_money is not None:
            _op_money = str(_op_money)

        if _op_user is not None:
            _op_user = _op_user.strip()

        if _op_depart is not None:
            _op_depart = _op_depart.strip()

        if _op_account in bank_info:
            true_depart = bank_info[_op_account]
            if true_depart != _op_depart:
                _my_key = str(i)    # 如果记账部门有误, 则把字典的key值类型设置为string, 原本为int
                output_data[_my_key] = [_op_user, _op_money, true_depart]
            else:
                output_data[i] = [_op_user, _op_money, _op_depart]
        else:
            err_sheet.cell(err_sheet_row, 2, i)
            err_sheet.cell(err_sheet_row, 3, '付方账号错误')
            err_sheet.cell(err_sheet_row, 7, _op_account)
            err_sheet_row += 1

    # 遍历[上传模板sheet], 获取 行号, 收方名称, 付款金额, 记账部门; 并向output_data做匹配
    for i in range(2, ori_sheet.max_row + 1):
        _money = ori_sheet.cell(i, 6).value
        _depart = ori_sheet.cell(i, 1).value
        _user = ori_sheet.cell(i, 3).value
        if _money is not None and _depart is not None and _user is not None:
            _money = str(_money)
            _depart = _depart.strip()
            _user = _user.strip()

            _compare_data = [_user, _money, _depart]
            output_value_list = output_data.values()

            if _compare_data in output_value_list:
                _my_index = output_value_list.index(_compare_data)
                _my_key = output_data.keys()[_my_index]
                if type(_my_key) is str:
                    err_sheet.cell(err_sheet_row, 1, i)
                    err_sheet.cell(err_sheet_row, 2, int(_my_key))
                    err_sheet.cell(err_sheet_row, 3, '记账部门有误')
                    err_sheet.cell(err_sheet_row, 4, _user)
                    err_sheet.cell(err_sheet_row, 5, float(_money)).number_format = numbers.FORMAT_NUMBER_00
                    err_sheet.cell(err_sheet_row, 6, _depart)
                    err_sheet_row += 1
                del output_data[_my_key]
            else:
                err_sheet.cell(err_sheet_row, 1, i)
                err_sheet.cell(err_sheet_row, 3, 'CannotMatch')
                err_sheet.cell(err_sheet_row, 4, _user)
                err_sheet.cell(err_sheet_row, 5, float(_money)).number_format = numbers.FORMAT_NUMBER_00
                err_sheet.cell(err_sheet_row, 6, _depart)
                err_sheet_row += 1

    if err_sheet_row == 2:
        wb.remove(wb.worksheets[sheet_num + 1])
    else:
        sheet_explain = '付方账号错误: 表示检索[银行导出表]时, "银行账号"没有在[银行明细]中找到\n' \
                        '记账部门有误: 表示[银行导出表]中, "银行账号" 和 "记账部门" 匹配不一致\n' \
                        'CannotMatch: 表示其他无法以[上传模板]中的条目 匹配 [银行导出]的各种情况\n' \
                        '             包括 收款人/付款不正确、付款银行未在资金池...等'
        err_sheet.cell(err_sheet_row + 2, 1, '本表解释')
        err_sheet.cell(err_sheet_row + 3, 1, sheet_explain)
    wb.save(filepath)
    return
