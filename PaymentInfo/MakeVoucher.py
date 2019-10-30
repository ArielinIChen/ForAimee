# encoding:utf-8
import os
import random
import xlwings
from openpyxl import Workbook

from common import output_info_redirect


def get_voucher_info(split_from_xl_path, pay_time, payee_data, paying_bank_data, logfile_path):
    # return data looks like {'voucher_data': {记账部门1-地区α: [数据A, 数据B....],
    #                                           记账部门2-地区β: [数据C, 数据D....], ......}
    #                         'error_data': [[u'错误原因', ....], [数据E, 数据F...], [数据G, 数据H...], ......]}
    # split_from_xl_path = u'E:\诺互银行\付款总计\mm.dd\mm.dd日款.xlsx'
    # pay_time is a string looks like YYYY-mm-dd
    # payee_data got from DependedInfo.get_payee_info
    # paying_bank_data got from DependedInfo.get_paying_bank_info

    pay_time = int(str(pay_time).replace('-', ''))
    app = xlwings.App(visible=True, add_book=False)
    wb = app.books.open(split_from_xl_path)
    # ws = wb.sheets[u'要付款明细']
    ws = wb.sheets['Sheet1']
    last_row_num = ws.used_range.last_cell.row

    output_info_redirect(u'要付款明细最大行号为: %s' % last_row_num, logfile_path)

    voucher_data = {}
    error_data = [[u'错误原因', u'出款银行', u'摘要信息', u'最终收款人', u'银行卡号', u'开户银行',
                   u'支付金额', u'单据编号', u'签收人', u'费用类型', u'出款日期', u'申请人']]

    for row_num in range(2, last_row_num + 1):
        making_depart = ws.range('A' + str(row_num)).value
        if not making_depart:
            # 这一行是空行
            continue
        else:
            payee_account = ws.range('D' + str(row_num)).value

            # 如果"收款人账号"或"制单部门"没有在相应的xxx_data中找到,
            # 那么就把对应的这一整行数据, append到xxx_not_found列表中
            this_line_data = ws.range('A' + str(row_num)).expand('right').value
            # type(this_line_data) == list
            if payee_account not in payee_data:
                output_info_redirect(u'没有在"收款人信息"中找到[收款账号]:%s, 将把付款明细第%s行记入错误信息...'
                                     % (payee_account, row_num), logfile_path)
                this_line_data.insert(0, u'收款人信息不存在')
                error_data.append(this_line_data)
            elif making_depart not in paying_bank_data:
                output_info_redirect(u'没有在"付款映射信息"中找到[记账部门]:%s, 将把付款明细第%s行记入错误信息...'
                                     % (making_depart, row_num), logfile_path)
                this_line_data.insert(0, u'制单部门不存在')
                error_data.append(this_line_data)
            else:
                _append_data = [
                    payee_account,                           # C  收款人帐号   -   总表D列
                    this_line_data[2],                       # D  收款人名称   -   总表C列|this_line_data index 2
                    this_line_data[4],                       # E  收方开户支行 -   总表E列|this_line_data index 4
                    payee_data[payee_account]['province'],   # F  收款人所在省 -   根据收款人账号, 向<打款资料库>查到相应 省
                    payee_data[payee_account]['city'],       # G  收款人所在市 -   根据收款人账号, 向<打款资料库>查到相应 市
                    '',                                      # H  收方邮件地址 -   空白
                    '',                                      # I  收方移动电话 -   空白
                    u'人民币',                                # J  币种        -   [人民币]
                    '',                                      # K  付款分行     -   空白
                    u'普通',                                  # L  结算方式    -   [普通]
                    '',                                      # M  业务种类     -   空白
                    paying_bank_data[making_depart][0],      # N  付方帐号     -   根据depart_name匹配<银行核对公式>的[银行明细]表
                    pay_time,                                # O  期望日       -   输入的日期, 格式为YYYYmmdd
                    '',                                      # P  期望时间     -   空白
                    this_line_data[6],                       # Q  用途        -   总表G列|this_line_data index 6
                    this_line_data[5],                       # R  金额        -   总表F列|this_line_data index 5
                    '',                                      # S  收方行号     -   空白
                    '',                                      # T  收方开户银行  -   空白
                    '',                                      # U  业务摘要     -   空白
                ]
                making_depart_region = paying_bank_data[making_depart][1]
                voucher_key = '-'.join([making_depart, making_depart_region])
                if voucher_key in voucher_data.keys():
                    voucher_data[voucher_key].append(_append_data)
                else:
                    voucher_data[voucher_key] = [_append_data]

    wb.close()
    app.quit()
    resp_data = {'voucher_data': voucher_data}
    if len(error_data) > 1:
        resp_data['error_data'] = error_data
    return resp_data


def create_voucher_file_use_xlwings(formwork_path, voucher_data, logfile_path):
    # formwork_path = u'E:\诺互银行\银行制单模板\模板汇总'
    # voucher_data got from .get_voucher_info
    collect_path = os.path.join(formwork_path, u'模板汇总')
    app = xlwings.App(visible=True, add_book=False)
    common_data = [u'业务参考号', u'收款人编号', u'收款人账号', u'收款人名称', u'收方开户支行', u'收款人所在省',
                   u'收款人所在市', u'收方邮件地址', u'收方移动电话', u'币种', u'付款分行', u'结算方式', u'业务种类',
                   u'付方帐号', u'期望日', u'期望时间', u'用途', u'金额', u'收方行号', u'收方开户银行', u'业务摘要']

    for key in voucher_data:
        filename = '.'.join([key, 'xlsx'])
        voucher_file_path = os.path.join(collect_path, filename)
        depart_data = voucher_data[key]
        wb = app.books.add()
        ws = wb.sheets['Sheet1']
        business_num = random.randint(1, 9) * 1000 + 1
        start_line = 7

        ws.range('A6').value = common_data
        for i in range(len(depart_data)):
            ws.range('A' + str(start_line)).value = [business_num, business_num]
            ws.range('C' + str(start_line)).value = depart_data[i]
            business_num += 1
            start_line += 1

        wb.save(voucher_file_path)
        wb.close()

    app.quit()
    return


def create_voucher_file_use_openpyxl(voucher_path, voucher_data, logfile_path):
    # voucher_path = u'E:\诺互银行\银行制单模板'
    # voucher_data got from .get_voucher_info
    common_data = [u'业务参考号', u'收款人编号', u'收款人账号', u'收款人名称', u'收方开户支行', u'收款人所在省',
                   u'收款人所在市', u'收方邮件地址', u'收方移动电话', u'币种', u'付款分行', u'结算方式', u'业务种类',
                   u'付方帐号', u'期望日', u'期望时间', u'用途', u'金额', u'收方行号', u'收方开户银行', u'业务摘要']

    for key in voucher_data:
        making_depart = key.split('-')[0]
        making_depart_region = key.split('-')[1]
        filename = '.'.join([making_depart, 'xlsx'])
        voucher_file_path = os.sep.join([voucher_path, making_depart_region, filename])
        depart_data = voucher_data[key]
        wb = Workbook()
        ws = wb.active
        for i in range(1, len(common_data) + 1):
            ws.cell(6, i, common_data[i - 1])

        business_num = random.randint(1, 9) * 1000 + 1
        start_line = 7

        for data in depart_data:
            for i in range(len(data)):
                ws.cell(start_line, i + 3, data[i])
            ws.cell(start_line, 1, business_num)
            ws.cell(start_line, 2, business_num)
            start_line += 1
            business_num += 1
        wb.save(voucher_file_path)
        wb.close()
    return
