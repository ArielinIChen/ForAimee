# encoding:utf-8
import os
import xlwings
from openpyxl import load_workbook

from common import output_info_redirect


def get_accountant_src_xl_data(accountant_src_xl_path, shop_depart_relation_data, date, logfile):
    # accountant_src_xl_path = u'E:\诺互银行\会计给到原件'
    # shop_depart_relation_data is a dict, got from function shop_depart_relationship
    # date looks like YYYY-MM-DD
    resp_dict = {}
    data_list = []
    no_depart_data_list = []
    xl_file_list = os.listdir(accountant_src_xl_path)
    for _xl_file in xl_file_list:
        if not _xl_file.endswith('.xlsx'):
            continue
        else:
            _xl_file_path = os.path.join(accountant_src_xl_path, _xl_file)
            _src_wb = load_workbook(_xl_file_path)
            _ws_names = _src_wb.sheetnames
            if 'Sheet1' in _ws_names:
                _src_ws = _src_wb['Sheet1']
            else:
                _src_ws = _src_wb[_ws_names[0]]
            _my_max_row = _src_ws.max_row
            _my_max_col = _src_ws.max_column
            # 从第二行开始遍历,
            # 如果这一行没有数据, 则pass;
            # 如果这一行有数据, 则判断 店名(第六列) 是否在shop_depart_relation_data的key_list中
            #   - 如果没有, 则把一整行数据记入no_depart_data_list
            #   - 如果有, 则摘取个别数据, 记入data_list
            for _row in range(2, _my_max_row + 1):
                if _src_ws.cell(_row, 2).value:
                    _my_shop = _src_ws.cell(_row, 6).value
                    if _my_shop not in shop_depart_relation_data:
                        text = u'%s 第%i行, 未找到[%s]的记账部门 \n' % (_xl_file, _row, _src_ws.cell(_row, 6).value)
                        output_info_redirect(text, logfile)
                        _no_depart_data = []
                        # 如果no_depart_data_list没有数据, 则向列表中加入标题行(即: 第一行)
                        if len(no_depart_data_list) == 0:
                            _data_row_list = [1, _row]
                        else:
                            _data_row_list = [_row]
                        for i in _data_row_list:
                            for _col in range(1, _my_max_col + 1):
                                _no_depart_data.append(_src_ws.cell(i, _col).value)
                            no_depart_data_list.append(_no_depart_data)
                        continue
                    else:
                        depart_name = shop_depart_relation_data[_my_shop]
                        summary = _src_ws.cell(_row, 12).value
                        payee_name = _src_ws.cell(_row, 8).value
                        payee_card_num = str(_src_ws.cell(_row, 11).value).replace(' ', '')
                        payee_bank = _src_ws.cell(_row, 10).value
                        amount = float(_src_ws.cell(_row, 9).value)
                        bill_num = _src_ws.cell(_row, 3).value
                        sign_by = _src_ws.cell(_row, 16).value
                        expense_type = _src_ws.cell(_row, 17).value
                        pay_at = date
                        applicant = _src_ws.cell(_row, 7).value
                        data_list.append([depart_name, summary, payee_name, payee_card_num, payee_bank, amount,
                                          bill_num, sign_by, expense_type, pay_at, applicant])
                else:
                    continue

    resp_dict['data_list'] = data_list
    if len(no_depart_data_list) > 0:
        resp_dict['no_depart_data_list'] = no_depart_data_list
    return resp_dict


def merge_payment_xl(total_xl, merge_data, logfile):
    # total_xl = u'E:\诺互银行\付款总计\month.day\month.day付款总计.xlsx'
    # accountant_src_xl_path = u'E:\诺互银行\会计给到原件'
    # merge_data is a list looks like [[depart_name1, summary1, .....], [depart_name2, summary2, ....], ....]
    app = xlwings.App(visible=True, add_book=False)
    total_wb = app.books.open(total_xl)
    payment_detail_ws = total_wb.sheets[u'要付款明细']
    # payment_detail_ws = total_wb.sheets['Sheet1']

    # 获取used_range, 根据最后一个cell找到最大行号
    sht_max_row_num = payment_detail_ws.used_range.last_cell.row

    output_info_redirect(u'要付款明细最大行号为: %s\n待合并的数据共: %s 行\n'
                         % (sht_max_row_num, len(merge_data)), logfile)

    depart_name_tag_dict = {}   # 记录depart_name及它所对应的最后一行号
    depart_name_tag_list = []   # 按顺序append depart_name, 当有插入行操作时, 此depart_name和它之后的所有depart_name 所对应的插入行号 + 1

    # 1. 遍历[要付款明细]sheet表的A列, 如果连续两个单元格内容不一样,
    #    则把 先一个单元格的内容 作为key, 行号 作为value 记入depart_name_tag_dict
    # 2. 并且按顺序将 先一个单元格的内容 记入depart_name_tag_list, 作为先后顺序, 在插入行时, 后面的depart_name对应的value + 1
    for i in range(2, sht_max_row_num + 1):
        depart_name_1 = payment_detail_ws.range('A' + str(i)).value
        depart_name_2 = payment_detail_ws.range('A' + str(i+1)).value
        if depart_name_1 != depart_name_2:
            depart_name_tag_dict[depart_name_1] = i
            depart_name_tag_list.append(depart_name_1)

    # 遍历会计原始数据
    for data in merge_data:
        # 找到所需插入的行号, 插入数据
        _depart_name = data[0]
        if _depart_name in depart_name_tag_dict:
            # 如果 日款表 中, 找到 记账部门, 直接添加
            insert_at = depart_name_tag_dict[_depart_name] + 1
            payment_detail_ws.api.Rows(insert_at).Insert()
            payment_detail_ws.range('A' + str(insert_at)).value = data
            # 此depart_name和它之后的所有depart_name 所对应的插入行号 + 1   #####特别备注, 包括它自己
            depart_index = depart_name_tag_list.index(_depart_name)
            for _name in depart_name_tag_list[depart_index:]:
                depart_name_tag_dict[_name] = depart_name_tag_dict[_name] + 1
        else:
            # # 如果 日款表 中, 没有找到 记账部门, 则往最后一行添加此条目; excel的最后一行空白行号 = sht_max_row_num + 1
            # insert_at = sht_max_row_num + 1
            # payment_detail_ws.range('A' + str(insert_at)).value = data
            # depart_name_tag_dict[_depart_name] = sht_max_row_num
            # depart_name_tag_list.append(_depart_name)
            # # 加在最后一行有点问题. 个别条目会隔非常多空行后, 添加在excel表最后
            #
            # 更改为: 在表格头部(第二行)插入 没有找到记账部门的付款信息
            #
            insert_at = 2
            payment_detail_ws.api.Rows(insert_at).Insert()
            payment_detail_ws.range('A' + str(insert_at)).value = data
            depart_name_tag_dict[_depart_name] = 2
            for _name in depart_name_tag_list[:]:
                depart_name_tag_dict[_name] = depart_name_tag_dict[_name] + 1

        sht_max_row_num += 1
        output_info_redirect(u'%s插入在表格当前的第%s行, 现在表格共%s行' % (data[6], insert_at, sht_max_row_num), logfile)

    # 保存并退出
    total_wb.save()
    total_wb.close()
    app.quit()
    return


# if __name__ == '__main__':
#     from DependedInfo import shop_depart_relationship
#     shop_bank = shop_depart_relationship(u'E:\诺互银行\银行信息资料\制单银行-门店对应表.xlsx')
#     total_excel = u'E:\诺互银行\付款总计\\10.17\\10.17付款总计.xlsx'
#     accountant_src_xl = u'E:\诺互银行\会计给到原件'
#     accountant_data = get_accountant_src_xl_data(accountant_src_xl, shop_bank, '2019-10-24')
#     print accountant_data
#     merge_payment_xl(total_excel, accountant_data)

