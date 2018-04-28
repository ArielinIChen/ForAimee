# encoding:utf-8
import xlrd
import openpyxl
import os
import shutil
import time
from common import copy_sheet, change_sheet_name, output_info_redirect
from common import _del_path


def core_method(dst_dir, err_dir, wt_logfile):
    final_echo = {'calc_err': [],
                  'map_err': [],
                  }
    shop_name_map_file_dict = {}
    dept_list = os.listdir(dst_dir)

    text = u'正在检查 利润表合并 文件夹...\n'
    output_info_redirect(text, wt_logfile)

    if u'利润表合并' not in dept_list:
        text = u'发现错误!\n' \
               u'%s 中没有找到 利润表合并 文件夹 请检查...\n' % dst_dir + \
               u'系统将退出! 请按 [回车键] 退出!\n'
        output_info_redirect(text, wt_logfile)
        raw_input()
        return

    merge_dir = os.path.join(dst_dir, u'利润表合并')
    if u'上海诺互利润表.xlsx' not in os.listdir(merge_dir):
        text = u'发现错误!\n' \
               u'%s 中没有找到 上海诺互利润表.xlsx 请检查...\n' % merge_dir + \
               u'系统将退出! 请按 [回车键] 退出!\n'
        output_info_redirect(text, wt_logfile)
        raw_input()
        return
    else:
        merge_file = os.path.join(merge_dir, u'上海诺互利润表.xlsx')
        merge_wb = openpyxl.load_workbook(filename=merge_file)
        merge_wb_shts = merge_wb.sheetnames

    dept_list.remove(u'利润表合并')
    merge_data = {}  # 存放门店的利润表数据
    text = u'检查完毕!\n' \
           u'正在检查 城市合并版 文件夹...'
    output_info_redirect(text, wt_logfile)

    combine_dir = os.path.join(dst_dir, u'城市合并版')
    _del_path(combine_dir, wt_logfile)
    # prime_dept_name = [u'北京诺互', u'成都诺互', u'广州景沪', u'广州诺互', u'海口诺互', u'昆明百当诺互',
    #                    u'南京诺互', u'青岛楷模', u'上海昌保', u'深圳深南', u'沈阳诺互', u'天津诺互',
    #                    u'武汉云景', u'西安来护']
    calc_err_dir = os.path.join(err_dir, u'不平的报表')
    merge_err_dir = os.path.join(err_dir, u'合并时出错的报表')
    # os.makedirs(calc_err_dir)

    text = u'开始处理利润表...\n'
    output_info_redirect(text, wt_logfile)

    for dept in dept_list:
        dept_dir = os.path.join(dst_dir, dept)
        shops_list = os.listdir(dept_dir)
        try:
            city_combine_filename = [x for x in shops_list if x.startswith(dept)][0]
        except IndexError:
            text = u'#########\n发现错误!\n' \
                   u'%s 文件夹中, 没有以 %s 开头的城市利润汇总表\n' \
                   u'#########' % (dept, dept)
            output_info_redirect(text, wt_logfile)
            continue
        shops_list.remove(city_combine_filename)
        city_combine_file = os.path.join(dept_dir, city_combine_filename)
        rb = xlrd.open_workbook(city_combine_file, formatting_info=True)
        total_profit = rb.sheet_by_index(0).cell(17, 1).value

        text = u'\n' + u'###' * 34 + u'\n#\n' + u'###' * 34 + u'\n' + \
               u'%s 在 汇总表中的净利润是: %s\n' % (dept, total_profit) + \
               u'**' * 20 + u'\n' + \
               u'准备修改汇总表 %s 的sheet名称...' % city_combine_filename
        output_info_redirect(text, wt_logfile)

        date_part = city_combine_filename.split('-')[-1]
        combine_sht_name = dept + u'-利润表-' + date_part.split('.xls')[0]
        change_sheet_name(city_combine_file, combine_sht_name, sht_num=0)

        cp_to_sht_num = 2
        from_sht_num = 1

        if len(shops_list) > 0:
            city_profit = 0.0
        else:
            city_profit = total_profit
        merge_shops_name = []
        month = int(date_part.split('.')[1])
        merge_col = month * 3 - 1
        for shop in shops_list:
            shop_file = os.path.join(dept_dir, shop)
            shop_name_tmp = shop.split(u'利润表')[0]
            shop_name = shop_name_tmp + u'店' if not shop_name_tmp.endswith(u'店') else shop_name_tmp
            cp_to_sht_name = shop_name + u'-利润表-' + date_part.split('.xls')[0]

            # 获取当月的实际净利润
            shop_rb = xlrd.open_workbook(shop_file, formatting_info=True)
            shop_profit = shop_rb.sheet_by_index(0).cell(17, 1).value
            city_profit += shop_profit

            text = u'**' * 20 + u'\n' + \
                   u'*开始编辑 %s\n' \
                   u'*加上 %s 的 %s 后, %s 的实际净利润是: %s\n' \
                   % (shop, shop_name, shop_profit, dept, city_profit) + \
                   u'**' * 20
            output_info_redirect(text, wt_logfile)

            # 获取当月的其他数据 写入字典merge_data
            shop_rs = shop_rb.sheet_by_index(0)
            merge_data[shop_name] = shop_rs.col_values(1, 1, 21)
            merge_shops_name.append(shop_name)
            shop_name_map_file_dict[shop_name] = {'name': shop, 'path': shop_file}

            # 调用common方法中的 - 复制sheet表函数
            copy_sheet(shop_file, city_combine_file, from_sht_num, cp_to_sht_num, cp_to_sht_name, wt_logfile=wt_logfile)

        total_profit = float(str(total_profit).decode('utf-8'))
        city_profit = float(str(city_profit).decode('utf-8'))

        text = u'@@' * 20 + u'\n' + \
               u'现在 %s 的总实际净利润是: %s\n' % (dept, city_profit) + \
               u'开始计算 %s 利润表B18单元格(净利润) 是否准确...\n' \
               u'汇总数额 %s VS 实际数额 %s\n' % (dept, total_profit, city_profit)
        output_info_redirect(text, wt_logfile)

        # 如果total数和实际数不一致, 那么报错,
        # 如果total数和实际数一致, 那么归档报表, 并把merge_data写入利润表合并
        if total_profit != city_profit:
            text = u'%s B18(净利润) 有错误, 将把该文件夹移动到 [不平的报表] 中...' % dept
            output_info_redirect(text, wt_logfile)

            os.path.exists(calc_err_dir) and 1 or os.mkdir(calc_err_dir)

            move_to_dir = os.path.join(calc_err_dir, dept)
            shutil.move(dept_dir, move_to_dir)

            final_echo['calc_err'].append(dept_dir)

            text = u'移动完毕!\n'
            output_info_redirect(text, wt_logfile)
        else:
            # 汇总 [利润表 - 城市合并版]到同一个文件夹
            name_part = dept + u'利润表'
            combine_file_name = '-'.join((name_part, date_part))
            combine_file = os.path.join(combine_dir, combine_file_name)

            text = u'准确无误!\n' \
                   u'准备将 汇总表 %s\n' \
                   u'复制到 城市合并版 %s\n' % (city_combine_file, combine_file)
            output_info_redirect(text, wt_logfile)

            shutil.copy(city_combine_file, combine_file)

            text = u'复制完成!\n\n' \
                   u'开始合并利润表...'
            output_info_redirect(text, wt_logfile)

            # 整理 利润表合并
            for merge_shop in merge_shops_name:
                try:
                    merge_sht_num = map((lambda y: merge_shop in y), merge_wb_shts).index(True)
                except ValueError:
                    text = u'没有在 %s 中找到 %s 的sheet名称, 或该名称不匹配\n' \
                           u'将把该文件 移动到 %s , 并跳过汇总过程...' \
                           % (u'上海诺互利润表.xlsx', merge_shop, merge_err_dir)
                    output_info_redirect(text, wt_logfile)

                    os.path.exists(merge_err_dir) and 1 or os.mkdir(merge_err_dir)
                    from_file = shop_name_map_file_dict[merge_shop]['path']
                    move_to_file = os.path.join(merge_err_dir, shop_name_map_file_dict[merge_shop]['name'])
                    shutil.move(from_file, move_to_file)
                    final_echo['map_err'].append(merge_shop)

                    text = u'移动完毕!\n'
                    output_info_redirect(text, wt_logfile)
                    continue
                else:
                    merge_ws = merge_wb[merge_wb_shts[merge_sht_num]]
                    data = merge_data[merge_shop]

                for i in range(len(data)):
                    fill_num = data[i]
                    if i == 0 and fill_num != 0:
                        merge_ws.cell(3, merge_col, fill_num)
                    elif i not in [10, 14, 16] and fill_num != 0:
                        merge_row = i + 5
                        merge_ws.cell(merge_row, merge_col, fill_num)
                merge_wb.save(merge_file)

                text = u'%s 利润表合并完毕' % merge_shop
                output_info_redirect(text, wt_logfile)

    merge_wb.close()

    text = u'\n' + u'###' * 20 + \
           u'\n利润表 Excel文件处理完毕.\n' \
           u'请在 %s 中获取处理后的Excel文件\n' % dst_dir
    output_info_redirect(text, wt_logfile)

    if os.listdir(err_dir):
        err_ori_xls_dir = os.path.join(err_dir, u'有错误的原始报表')
        if os.path.exists(err_ori_xls_dir):
            for i in os.listdir(err_ori_xls_dir):
                print i,
                i = i.encode('utf-8')
                with open(wt_logfile, 'a+') as write_file:
                    write_file.write(i + '\n')
            print ('')
            text = u'原始报表有误, 请在 %s 获取相关报表\n' % err_ori_xls_dir
            output_info_redirect(text, wt_logfile)

        if len(final_echo['calc_err']) > 0:
            for i in final_echo['calc_err']:
                print i,
                i = i.encode('utf-8')
                with open(wt_logfile, 'a+') as write_file:
                    write_file.write(i + '\n')
            print ('')
            text = u'报表不平, 请在 %s 获取相关报表\n' % calc_err_dir
            output_info_redirect(text, wt_logfile)

        if len(final_echo['map_err']) > 0:
            for i in final_echo['map_err']:
                print i,
                i = i.encode('utf-8')
                with open(wt_logfile, 'a+') as write_file:
                    write_file.write(i + '\n')
            print ('')
            text = u'在合并报表时, 没有匹配到相应sheet, 请在 %s 获取相关报表\n' % merge_err_dir
            output_info_redirect(text, wt_logfile)
    else:
        text = u'并且, 没有找到出错的报表! Good Job~\n'
        output_info_redirect(text, wt_logfile)
        os.removedirs(err_dir)

    text = u'###' * 20 + \
           u'\n请按 [回车键] 退出...'
    output_info_redirect(text, wt_logfile)
    raw_input()
    time.sleep(1)
    return
